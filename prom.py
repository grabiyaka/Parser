import openpyxl
from datetime import datetime

# Укажите путь к вашему XLSX файлу
xlsx_file_path = "xlsx/ExportProducts_Prom.xlsx"

# Открываем файл
workbook = openpyxl.load_workbook(xlsx_file_path)

# Словарь для хранения ссылок на изображения
products = []

# Перебираем все листы в файле
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    
    # Получаем названия колонок (заголовки)
    headers = [cell.value for cell in sheet[1]]

    # Перебираем строки в текущем листе
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = {
            'productId': row[headers.index('Код_товара')],
            'image': row[headers.index('Ссылка_изображения')].split(',')[0] if 'Ссылка_изображения' in headers and headers.index('Ссылка_изображения') < len(row) and row[headers.index('Ссылка_изображения')] is not None else None,
            'description': row[headers.index('Описание')],
            'name': row[headers.index('Название_позиции')],
            'price': row[headers.index('Цена')],
            'currencyId': row[headers.index('Валюта')],
            'vendor': row[headers.index('Производитель')],
            'vendorId': row[headers.index('Номер_группы')]
        }
        products.append(item)
        
# Закрываем файл
workbook.close()

def generate_xml_t(products):
    output = '<?xml version="1.0" encoding="UTF-8"?>\n' \
             '<yml_catalog date="' + datetime.now().strftime('%Y-%m-%d %H:%M') + '">\n' \
             '    <shop>\n' \
             '        <currencies>\n' \
             '            <currency id="UAH" rate="1"/>\n' \
             '        </currencies>\n' \
             '        <categories>\n' \
             '            <category id="5875416">Пocyда</category>\n' \
             '            <category id="5926324" parentId="5875416">Kуxонная посудa</category>\n' \
             '            <category id="5947705" parentId="5926324">Кофейные турки</category>\n' \
             '            <category id="5940226" parentId="5926324">Кастрюли</category>\n' \
             '        </categories>\n' \
             '        <offers>'

    for index, product in enumerate(products):
        output += '<offer id="' + str(product['productId']) + '" available="true">\n' \
                  '        <name>' + str(product['name']) + '</name>\n' \
                  '        <price>' + str(product['price']) + '</price>\n' \
                  '        <currencyId>' + str(product['currencyId']) + '</currencyId>\n' \
                  '        <vendor>' + str(product['vendor']) + '</vendor>\n' \
                  '        <vendorCode>' + str(product['vendorId']) + '</vendorCode>\n' \
                  '        <description><![CDATA[' + str(product['description']) + ']]></description>\n' \
                  '        <categoryId></categoryId>\n' \
                  '        <picture>' + str(product['image']) + '</picture>\n' \
                  '</offer>'

    output += '\n' \
              '        </offers>\n' \
              '    </shop>\n' \
              '</yml_catalog>'
    return output


def save_to_xml(data, file_path):
    with open(file_path, 'w', encoding='utf-8') as file:
        file.write(data)
result = generate_xml_t(products)
save_to_xml(result, 'xml/ExportProducts_Prom.xml')